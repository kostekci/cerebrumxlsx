import pandas
import time
import win32clipboard
from openpyxl import load_workbook

last_text=''

d1 = {'Words': ['1']}
df1 = pandas.DataFrame(data=d1)

while True:
    
    # get clipboard data
    win32clipboard.OpenClipboard()
    text = win32clipboard.GetClipboardData()
    win32clipboard.CloseClipboard()
    
    if text != last_text:
        
        last_text=text
        
        #append to file instead of write
        book = load_workbook('CheckedWords.xlsx')
        writer = pandas.ExcelWriter('CheckedWords.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        df1['Words']=text
        for sheetname in writer.sheets:
            df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

        writer.save()
    time.sleep(0.5)